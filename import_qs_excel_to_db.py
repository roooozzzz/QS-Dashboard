import os
import sqlite3
import openpyxl

# 国家到地区映射（部分示例，后续可补充完整）
country_region_map = {
    # Africa
    'South Africa': 'Africa', 'Egypt': 'Africa', 'Nigeria': 'Africa', 'Kenya': 'Africa', 'Morocco': 'Africa',
    'Ethiopia': 'Africa', 'Ghana': 'Africa', 'Libya': 'Africa', 'Sudan': 'Africa', 'Tunisia': 'Africa', 'Uganda': 'Africa',
    # Americas
    'United States': 'Americas', 'Canada': 'Americas', 'Brazil': 'Americas', 'Mexico': 'Americas', 'Argentina': 'Americas',
    'Bolivia': 'Americas', 'Chile': 'Americas', 'Colombia': 'Americas', 'Costa Rica': 'Americas', 'Cuba': 'Americas',
    'Dominican Republic': 'Americas', 'Ecuador': 'Americas', 'Guatemala': 'Americas', 'Honduras': 'Americas',
    'Panama': 'Americas', 'Paraguay': 'Americas', 'Peru': 'Americas', 'Puerto Rico': 'Americas', 'Uruguay': 'Americas',
    'Venezuela': 'Americas', 'Venezuela (Bolivarian Republic of)': 'Americas',
    # Asia
    'China': 'Asia', 'Japan': 'Asia', 'South Korea': 'Asia', 'Republic of Korea': 'Asia', 'India': 'Asia', 'Singapore': 'Asia', 'Hong Kong SAR': 'Asia', 'Taiwan': 'Asia', 'Malaysia': 'Asia', 'Thailand': 'Asia', 'Indonesia': 'Asia', 'Pakistan': 'Asia', 'Bangladesh': 'Asia', 'Vietnam': 'Asia', 'Viet Nam': 'Asia', 'Philippines': 'Asia',
    'Armenia': 'Asia', 'Azerbaijan': 'Asia', 'Bahrain': 'Asia', 'Brunei': 'Asia', 'Brunei Darussalam': 'Asia', 'China (Mainland)': 'Asia', 'Cyprus': 'Asia', 'Georgia': 'Asia', 'Iran (Islamic Republic of)': 'Asia', 'Iran, Islamic Republic of': 'Asia', 'Iraq': 'Asia', 'Israel': 'Asia', 'Jordan': 'Asia', 'Kazakhstan': 'Asia', 'Kuwait': 'Asia', 'Kyrgyzstan': 'Asia', 'Lebanon': 'Asia', 'Macao SAR, China': 'Asia', 'Northern Cyprus': 'Asia', 'Oman': 'Asia', 'Palestine': 'Asia', 'Palestinian Territory, Occupied': 'Asia', 'Qatar': 'Asia', 'Saudi Arabia': 'Asia', 'Sri Lanka': 'Asia', 'Syrian Arab Republic': 'Asia', 'Türkiye': 'Asia', 'United Arab Emirates': 'Asia', 'Uzbekistan': 'Asia',
    # Europe
    'United Kingdom': 'Europe', 'Germany': 'Europe', 'France': 'Europe', 'Italy': 'Europe', 'Spain': 'Europe', 'Russia': 'Europe', 'Netherlands': 'Europe', 'Sweden': 'Europe', 'Switzerland': 'Europe', 'Belgium': 'Europe', 'Denmark': 'Europe', 'Finland': 'Europe', 'Norway': 'Europe', 'Poland': 'Europe', 'Austria': 'Europe', 'Ireland': 'Europe', 'Portugal': 'Europe', 'Czech Republic': 'Europe', 'Czechia': 'Europe', 'Greece': 'Europe', 'Hungary': 'Europe', 'Romania': 'Europe', 'Turkey': 'Europe', 'Ukraine': 'Europe', 'Serbia': 'Europe', 'Croatia': 'Europe', 'Slovakia': 'Europe', 'Slovenia': 'Europe', 'Bulgaria': 'Europe', 'Estonia': 'Europe', 'Lithuania': 'Europe', 'Latvia': 'Europe', 'Luxembourg': 'Europe', 'Iceland': 'Europe', 'Belarus': 'Europe', 'Bosnia and Herzegovina': 'Europe', 'Malta': 'Europe', 'Russian Federation': 'Europe',
    # Oceania
    'Australia': 'Oceania', 'New Zealand': 'Oceania',
}
def get_region(country):
    if not country:
        return None
    # 兼容部分国家名不同写法
    if country in country_region_map:
        return country_region_map[country]
    # 处理部分特殊情况
    if 'Hong Kong' in country:
        return 'Asia'
    if 'Macau' in country:
        return 'Asia'
    if 'Taiwan' in country:
        return 'Asia'
    if 'United States' in country or 'USA' in country:
        return 'Americas'
    if 'UK' in country or 'Britain' in country:
        return 'Europe'
    return None

data_dir = 'data'
db_path = os.path.join(data_dir, 'qs_rankings.db')
excel_files = [
    '2022QSRankings.xlsx',
    '2023QSRankings.xlsx',
    '2024QSRankings.xlsx',
    '2025QSRankings.xlsx',
    '2026QSRankings.xlsx',
]

# 按 @表头 要求的顺序
columns = [
    'RANK', 'NAME', 'COUNTRY', 'YEAR', 'REGION', 'TOTAL_SCORE',
    'AR_SCORE', 'AR_RANK',
    'ER_SCORE', 'ER_RANK',
    'FSR_SCORE', 'FSR_RANK',
    'CPF_SCORE', 'CPF_RANK',
    'IFR_SCORE', 'IFR_RANK',
    'ISR_SCORE', 'ISR_RANK',
    'ISD_SCORE', 'ISD_RANK',
    'IRN_SCORE', 'IRN_RANK',
    'EO_SCORE', 'EO_RANK',
    'SUS_SCORE', 'SUS_RANK',
]

# 各年份缺失的字段
missing_fields = {
    2022: ['EO_SCORE', 'EO_RANK', 'SUS_SCORE', 'SUS_RANK', 'ISD_SCORE', 'ISD_RANK'],
    2023: ['SUS_SCORE', 'SUS_RANK', 'ISD_SCORE', 'ISD_RANK'],
    2024: ['ISD_SCORE', 'ISD_RANK'],
    2025: ['ISD_SCORE', 'ISD_RANK'],
    2026: [],
}

def create_table(conn):
    cur = conn.cursor()
    cur.execute('DROP TABLE IF EXISTS qs_rankings')
    cur.execute(f'''
        CREATE TABLE qs_rankings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            RANK TEXT,
            NAME TEXT,
            COUNTRY TEXT,
            YEAR INTEGER,
            REGION TEXT,
            TOTAL_SCORE TEXT,
            AR_SCORE TEXT,
            AR_RANK TEXT,
            ER_SCORE TEXT,
            ER_RANK TEXT,
            FSR_SCORE TEXT,
            FSR_RANK TEXT,
            CPF_SCORE TEXT,
            CPF_RANK TEXT,
            IFR_SCORE TEXT,
            IFR_RANK TEXT,
            ISR_SCORE TEXT,
            ISR_RANK TEXT,
            ISD_SCORE TEXT,
            ISD_RANK TEXT,
            IRN_SCORE TEXT,
            IRN_RANK TEXT,
            EO_SCORE TEXT,
            EO_RANK TEXT,
            SUS_SCORE TEXT,
            SUS_RANK TEXT
        )
    ''')
    conn.commit()

def import_excel_to_db():
    conn = sqlite3.connect(db_path)
    create_table(conn)
    cur = conn.cursor()
    region_missing_count = 0
    for file in excel_files:
        year = int(file[:4])
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        seen = set()  # 用于去重
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(header, row))
            # 跳过无效行
            if not row_dict.get('NAME') or not row_dict.get('COUNTRY'):
                continue
            key = (row_dict.get('NAME'), year)
            if key in seen:
                continue
            seen.add(key)
            country = row_dict.get('COUNTRY')
            region = get_region(country)
            if not region:
                region_missing_count += 1
            data = {
                'RANK': row_dict.get('RANK'),
                'NAME': row_dict.get('NAME'),
                'COUNTRY': country,
                'YEAR': year,
                'REGION': region,
                'TOTAL_SCORE': row_dict.get('TOTAL SCORE'),
                'AR_SCORE': row_dict.get('AR SCORE'),
                'AR_RANK': row_dict.get('AR RANK'),
                'ER_SCORE': row_dict.get('ER SCORE'),
                'ER_RANK': row_dict.get('ER RANK'),
                'FSR_SCORE': row_dict.get('FSR SCORE'),
                'FSR_RANK': row_dict.get('FSR RANK'),
                'CPF_SCORE': row_dict.get('CPF SCORE'),
                'CPF_RANK': row_dict.get('CPF RANK'),
                'IFR_SCORE': row_dict.get('IFR SCORE'),
                'IFR_RANK': row_dict.get('IFR RANK'),
                'ISR_SCORE': row_dict.get('ISR SCORE'),
                'ISR_RANK': row_dict.get('ISR RANK'),
                'ISD_SCORE': row_dict.get('ISD SCORE'),
                'ISD_RANK': row_dict.get('ISD RANK'),
                'IRN_SCORE': row_dict.get('IRN SCORE'),
                'IRN_RANK': row_dict.get('IRN RANK'),
                'EO_SCORE': row_dict.get('EO SCORE'),
                'EO_RANK': row_dict.get('EO RANK'),
                'SUS_SCORE': row_dict.get('SUS SCORE'),
                'SUS_RANK': row_dict.get('SUS RANK'),
            }
            for field in missing_fields[year]:
                data[field] = None
            values = [data[col] for col in columns]
            cur.execute(
                "INSERT INTO qs_rankings (RANK, NAME, COUNTRY, YEAR, REGION, TOTAL_SCORE, AR_SCORE, AR_RANK, ER_SCORE, ER_RANK, FSR_SCORE, FSR_RANK, CPF_SCORE, CPF_RANK, IFR_SCORE, IFR_RANK, ISR_SCORE, ISR_RANK, ISD_SCORE, ISD_RANK, IRN_SCORE, IRN_RANK, EO_SCORE, EO_RANK, SUS_SCORE, SUS_RANK) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                values
            )
    conn.commit()
    conn.close()
    print(f'未能归类地区的高校数量: {region_missing_count}')

if __name__ == '__main__':
    import_excel_to_db()
    print('数据导入完成！') 